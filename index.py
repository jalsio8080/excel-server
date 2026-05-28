from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
import io
import os
from datetime import date

app = Flask(__name__)
CORS(app)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'template.xlsx')

def set_cell(ws, coord, value):
    """병합 셀 구조를 유지하면서 최상단 좌측 셀에 값 입력"""
    if value is None or value == '':
        return
    for merge in ws.merged_cells.ranges:
        if coord in merge:
            ws.cell(row=merge.min_row, column=merge.min_col).value = value
            return
    ws[coord].value = value

@app.route('/', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'message': '성공K 엑셀 서버 정상 동작 중'})

@app.route('/generate-excel', methods=['POST', 'OPTIONS'])
def generate_excel():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        p = request.get_json()
        if not p:
            return jsonify({'error': '데이터가 없습니다'}), 400

        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active

        today = date.today()
        today_str = f"{today.year}년 {today.month}월 {today.day}일"

        # ── 기본 정보 ──
        set_cell(ws, 'I2', today_str)                           # 상담일
        set_cell(ws, 'D2', p.get('bizName', ''))                # 업체명
        set_cell(ws, 'D3', p.get('bizNo', ''))                  # 사업자등록번호
        set_cell(ws, 'I3', p.get('corpNo', ''))                 # 법인등록번호
        set_cell(ws, 'D4', p.get('industry', ''))               # 업종
        set_cell(ws, 'I4', p.get('openDate', ''))               # 개업연월일
        set_cell(ws, 'D5', p.get('bizDetail', ''))              # 주 생산품목

        # 기계설비
        machine = p.get('machine', '무')
        if machine == '유':
            set_cell(ws, 'I5', '유 / 내용 : ' + p.get('machineDetail', ''))
        else:
            set_cell(ws, 'I5', '무')

        # ── 재무 정보 ──
        rev_now = str(p['rev2025']) + '만원' if p.get('rev2025') else p.get('revNow', '')
        if rev_now:
            set_cell(ws, 'E6', rev_now)                         # 현재 매출

        credit_parts = []
        if p.get('creditNice'): credit_parts.append(str(p['creditNice']))
        if p.get('creditKcb'):  credit_parts.append(str(p['creditKcb']))
        if credit_parts:
            set_cell(ws, 'I6', ' / '.join(credit_parts))        # 신용점수

        # 매출 (이 양식: B7=2022, F7=2023, I7=2024)
        if p.get('rev2022'): set_cell(ws, 'C7', str(p['rev2022']) + '만원')
        if p.get('rev2023'): set_cell(ws, 'G7', str(p['rev2023']) + '만원')
        if p.get('rev2024'): set_cell(ws, 'J7', str(p['rev2024']) + '만원')

        # ── 대출 현황 ──
        loan_list = p.get('loanList', [])
        if loan_list:
            loan_txt = ' / '.join([f"{l.get('name','')} {l.get('amount','')}만원" for l in loan_list])
            set_cell(ws, 'E8', loan_txt)                        # 기대출 운전

        # ── 필요자금 ──
        if p.get('needAmount'):
            set_cell(ws, 'E10', str(p['needAmount']) + '만원')  # 필요자금 운전
        if p.get('fundPurpose'):
            set_cell(ws, 'H10', p['fundPurpose'])               # 용도

        # ── 주소 ──
        if p.get('address'):
            set_cell(ws, 'D12', '사업자등록증상 주소 : ' + p['address'])
        set_cell(ws, 'I12', '임차')
        deposit = p.get('depositAmt', 0)
        monthly = p.get('monthlyRent', 0)
        if deposit or monthly:
            set_cell(ws, 'J12', f"{deposit or '-'}만 / {monthly or '-'}만")

        # ── 대표자 정보 ──
        set_cell(ws, 'D15', p.get('ceoName', ''))               # 성명
        set_cell(ws, 'D17', p.get('ceoPhone', ''))              # 연락처
        age_str = '청년대상 여부 (해당)' if p.get('ceoAgeGroup') == '청년' else '청년대상 여부 (비해당)'
        set_cell(ws, 'D18', age_str)

        # ── 직원수 ──
        if p.get('employees') is not None:
            set_cell(ws, 'I20', f"상시근로자 {p['employees']}명 (4대보험가입자   명)")

        # ── 여성기업 ──
        if p.get('isWoman'):
            set_cell(ws, 'I26', '유   /   무')

        # ── 비고 ──
        bigo_parts = []
        if p.get('specialNote'):    bigo_parts.append(p['specialNote'])
        if p.get('missingFields'):  bigo_parts.append('[추가 확인 필요] ' + ', '.join(p['missingFields']))
        if bigo_parts:
            set_cell(ws, 'C31', '\n'.join(bigo_parts))

        # ── 가점 ──
        open_year = p.get('openYear')
        if open_year and (today.year - int(open_year)) < 7:
            set_cell(ws, 'Q34', '해당')                         # 31번: 창업 7년 미만
        if p.get('isWoman'):
            set_cell(ws, 'Q8', '해당')                          # 6번: 여성기업

        # 버퍼에 저장 후 반환
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        biz_name = p.get('bizName', '고객')
        filename = f"{biz_name}_기업정보및가점평가표_{today_str.replace(' ', '')}.xlsx"

        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
